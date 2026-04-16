import openpyxl as xl # Обработка файла xlsx
import urllib # Работают вместе 
import requests # Работают вместе 
import timeit # Чекаем скорость работы
import json # Обработка формата данных
import os # Хранение данных

url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTOrEM-h36wb2LmpE3nF6f5tQMKVrBxWSZceBa7Jl5BZd5VaAt3GsHBLefgq9VV8RMDDb0FQEWNQol-/pub?output=xlsx"
# F1 - число
# I1 - день недели
current_dir = os.path.dirname(os.path.abspath(__file__))
if not os.path.exists(os.path.join(current_dir, "days")):
    os.makedirs(os.path.join(current_dir, "days"))

def save_sheet(data: dict, filename: str):
    with open(f"{current_dir}/days/{filename}.json", 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def read_sheet(filename: str):
    with open(f"{current_dir}/days/{filename}", 'r', encoding='utf-8') as f:
        return json.load(f)

def read_sheets():
    sheets = []
    for filename in os.listdir(f"{current_dir}/days"):
        if filename.endswith(".json"):
            sheets.append({
                filename.rstrip('.json'): read_sheet(filename),
                'date': filename.split()[0],
                'day': filename.split()[1].rstrip('.json')
                })
    return sheets

def update_tables(url=url):
    if url:
        try:
            urllib.request.urlretrieve(url, "tables.xlsx")
        except:
            print("Ошибка скачивания таблицы")
    
# ДАЖЕ КОММЕНТЫ ДОБАВИЛ, АРТЕМ НЕ ТУПИ
def parsing(url=None, sheet=None, is_save=False): 
    """Функция для извлечения данных из таблицы с расписанием уолледжа подмосковье\n
    Keyword arguments:\n
    ulr: Имеет приоритет над sheet, скачивает таблицу и берёт активный sheet.\n
    sheet: Обрабатывает конкретный sheet. Заранее инициализировать в переменную и передать в функцию.\n
    is_save: Если 'True', сохранияет обработанную таблицу в формате '.json' в папку 'jsons'.\n
    Return: Список групп.
    """
    
    if url:
        try:
            urllib.request.urlretrieve(url, "tables.xlsx")
            wb = xl.load_workbook("tables.xlsx", read_only=True)
            sheet = wb.active
        except:
            print("Ошибка скачавания таблицы")
        print("Данные обновлены")
        
    groups = []
    row_id = 0
    move = [0, 0, 0, 0]
    if sheet["A21"].value:
        move = [0, 3, 6, 9]
    for row in sheet.iter_rows():
        row_id += 1
        col_id = 0
        for col in row:
            col_id += 1
            if row_id == 3 or row_id == 24 + move[1] or row_id == 45 + move[2] or row_id == 66 + move[3]: # Строки с названиями групп
                if col_id > 2:
                    if col.value:
                        if row_id == 3:
                            curse = 1
                        if row_id == 24 + move[1]:
                            curse = 2
                        if row_id == 45 + move[2]:
                            curse = 3
                        if row_id == 66 + move[3]:
                            curse = 4
                        groups.append({
                            "id": col_id,
                            "name": str(col.value),
                            "curse": curse,
                            '1': None,
                            '2': None,
                            '3': None,
                            '4': None,
                            '5': None,
                            '6': None
                        })
    iterations = 5
    if move[1] != 0:
        iterations = 6
    for group in groups:
        if group["curse"] == 1:
            row = 3
        if group["curse"] == 2:
            row = 24
        if group["curse"] == 3:
            row = 45
        if group["curse"] == 4:
            row = 66
        for i in range(iterations):
            group[str(i+1)] = {
                "para": sheet.cell(row=row + i*3 + 3 + move[group["curse"]-1], column=group["id"]).value,
                "teacher": sheet.cell(row=row + i*3 + 4 + move[group["curse"]-1], column=group["id"]).value,
                }
    
    # Сохранение данных
    if is_save:
        date1 = sheet['F1'].value
        if "января" in sheet.title:
            date2 = 0.01
        elif "февраля" in sheet.title:
            date2 = 0.02
        elif "марта" in sheet.title:
            date2 = 0.03
        elif "апреля" in sheet.title:
            date2 = 0.04
        elif "мая" in sheet.title:
            date2 = 0.05
        elif "июня" in sheet.title:
            date2 = 0.06
        elif "июля" in sheet.title:
            date2 = 0.07
        elif "августа" in sheet.title:
            date2 = 0.08
        elif "сентября" in sheet.title:
            date2 = 0.09
        elif "октября" in sheet.title:
            date2 = 0.10
        elif "ноября" in sheet.title:
            date2 = 0.11
        elif "декабря" in sheet.title:
            date2 = 0.12
        save_sheet(data=groups, filename=f'{date1 + date2} {sheet["I1"].value}')  
    
    # Возвращает списко групп студентов
    return groups    

def test_url():
    for i in parsing(url=url):
        print(f"""
| Группа: {i["name"]} | Курс: {i["curse"]} |
|1| {i['1']}
|2| {i['2']}
|3| {i['3']}
|4| {i['4']}
|5| {i['5']}
|6| {i['6']}
""")
    print("Время обработки таблицы: ", timeit.timeit(lambda: parsing(url), number=1))

def test_sheets():
    wb = xl.load_workbook("tables.xlsx", read_only=True)
    sheets = wb.sheetnames
    for sheet in sheets:
        print(f"// ТАБЛИЦА {sheet} //")
        for i in parsing(sheet=wb[sheet]):
            print(f"""
| Группа: {i["name"]} | Курс: {i["curse"]} |
|1| {i['1']}
|2| {i['2']}
|3| {i['3']}
|4| {i['4']}
|5| {i['5']}
|6| {i['6']}
        """)

def test_save():
    wb = xl.load_workbook("tables.xlsx", read_only=True)
    sheets = wb.sheetnames
    for sheet in sheets:
        print(f"// ТАБЛИЦА {sheet} ОБРАБАТЫВАЕТСЯ //")
        parsing(sheet=wb[sheet], is_save=True)
        print(f"// ТАБЛИЦА {sheet} СОХРАНЕНА //")
        print()
        
if __name__ == "__main__":
    # test_url()
    # test_sheets()
    test_save()
    # update_tables()
    pass